import os
import shutil
import threading
from io import BytesIO
from kivy.uix.screenmanager import Screen
from kivy.clock import Clock
from kivy.utils import platform
from kivy.logger import Logger
from core.config import LAYOUT_PRESETS, OUTPUT_DIR
from core.preset_loader import PresetLoader
from core.excel_processor import ExcelProcessor
from core.traffic_logger import TrafficLogger
from core.utils import (
    convert_date_to_english,
    generate_filename,
    sort_photos_by_datetime,
    get_output_path,
)


class HomeScreen(Screen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.preset_loader = PresetLoader()
        self.traffic = TrafficLogger()
        self.photo_paths = []
        self.output_path = None
        self.output_bytes = None
        self.sheet_names = []
        self.temp_dir = None
        self._manual_template_bytes = None
        self._temp_counter = 0

    def on_enter(self):
        self.traffic.sync_on_startup()
        self._init_temp()
        self._setup_template_dropdown()
        self._setup_layout_dropdown()

    def _init_temp(self):
        if platform == "android":
            try:
                from jnius import autoclass
                PythonActivity = autoclass("org.kivy.android.PythonActivity")
                ctx = PythonActivity.mActivity
                cache_dir = ctx.getCacheDir().getAbsolutePath()
                self.temp_dir = os.path.join(cache_dir, "photos")
            except Exception:
                self.temp_dir = os.path.join(
                    os.path.dirname(OUTPUT_DIR), "temp_photos"
                )
        else:
            self.temp_dir = os.path.join(
                os.path.dirname(OUTPUT_DIR), "temp_photos"
            )
        os.makedirs(self.temp_dir, exist_ok=True)
        Logger.info(f"QC Teazzi: temp_dir = {self.temp_dir}")

    # ── TEMPLATE DROPDOWN ──
    def _setup_template_dropdown(self):
        options = self.preset_loader.get_dropdown_options()
        self.ids.dd_template.values = options
        cached = self.preset_loader.list_cached()
        if cached:
            self.ids.dd_template.text = f"{cached[0]} ✓"
            self._on_template_change(cached[0])
        else:
            self.ids.dd_template.text = "Upload Manual"

    def _on_template_selected(self, value):
        real = self.preset_loader.get_real_filename(value)
        if real == "Upload Manual":
            self._pick_template_file()
            return
        if real == "─────────────":
            return
        self._on_template_change(real)

    def _on_template_change(self, filename):
        self.sheet_names = self.preset_loader.get_sheet_names(filename)
        self.ids.dd_sheet.values = self.sheet_names
        if self.sheet_names:
            default = (
                "ATTACHMENT"
                if "ATTACHMENT" in self.sheet_names
                else self.sheet_names[0]
            )
            self.ids.dd_sheet.text = default
            Logger.info(f"QC Teazzi: Sheet selected = {default}")

    def _pick_template_file(self):
        try:
            from plyer import filechooser
            filechooser.open_file(
                on_selection=self._on_template_picked,
                filters=["*.xlsx"],
            )
        except Exception as e:
            Logger.error(f"QC Teazzi: Template picker error: {e}")
            self.ids.lbl_status.text = f"[color=ff6b6b]Gagal buka file picker: {e}[/color]"

    def _on_template_picked(self, selection):
        if not selection:
            return
        path = str(selection[0])
        Logger.info(f"QC Teazzi: Template picked = {path}")

        # Android returns content:// URI, need to copy to local
        local_path = self._resolve_to_local(path, prefix="template")
        if not local_path or not os.path.exists(local_path):
            self.ids.lbl_status.text = "[color=ff6b6b]Gagal membaca template[/color]"
            return

        try:
            from openpyxl import load_workbook
            wb = load_workbook(local_path, read_only=True)
            self.sheet_names = wb.sheetnames
            wb.close()
            with open(local_path, "rb") as f:
                self._manual_template_bytes = BytesIO(f.read())
            self.ids.dd_template.text = os.path.basename(path)
            self.ids.dd_sheet.values = self.sheet_names
            if self.sheet_names:
                self.ids.dd_sheet.text = self.sheet_names[0]
            Logger.info(f"QC Teazzi: Manual template loaded, sheets={self.sheet_names}")
        except Exception as e:
            Logger.error(f"QC Teazzi: Template read error: {e}")
            self.ids.lbl_status.text = f"[color=ff6b6b]Error baca template: {e}[/color]"

    # ── LAYOUT DROPDOWN ──
    def _setup_layout_dropdown(self):
        self.ids.dd_layout.values = list(LAYOUT_PRESETS.keys())
        self.ids.dd_layout.text = "LGJA"
        self._update_layout_info("LGJA")

    def _on_layout_selected(self, value):
        self._update_layout_info(value)

    def _update_layout_info(self, name):
        if name in LAYOUT_PRESETS:
            p = LAYOUT_PRESETS[name]
            rows = len(p["rows"])
            cols = len(p["cols"])
            self.ids.lbl_layout_info.text = (
                f"{rows} baris x {cols} kolom | {p['img_w']} x {p['img_h']} cm"
            )
        else:
            self.ids.lbl_layout_info.text = ""

    # ── FOTO PICKER ──
    def pick_photos(self):
        try:
            from plyer import filechooser
            filechooser.open_file(
                on_selection=self._on_photos_picked,
                multiple=True,
            )
        except Exception as e:
            Logger.error(f"QC Teazzi: Photo picker error: {e}")
            self.ids.lbl_status.text = f"[color=ff6b6b]Gagal buka gallery: {e}[/color]"

    def _on_photos_picked(self, selection):
        if not selection:
            return

        Logger.info(f"QC Teazzi: {len(selection)} photos selected")
        paths = []
        for item in selection:
            local_path = self._resolve_to_local(str(item), prefix="photo")
            if local_path and os.path.exists(local_path):
                paths.append(local_path)
                Logger.info(f"QC Teazzi: Resolved photo = {local_path}")
            else:
                Logger.warning(f"QC Teazzi: Failed to resolve = {item}")

        self.photo_paths = sort_photos_by_datetime(paths)
        self.ids.lbl_photo_count.text = f"{len(self.photo_paths)} foto dipilih"

    # ── RESOLVE URI TO LOCAL FILE ──
    def _resolve_to_local(self, uri_or_path, prefix="file"):
        """Convert any path/URI to a readable local file path."""
        path = str(uri_or_path)

        # If it's already a normal file path
        if not path.startswith("content://") and os.path.exists(path):
            return path

        if platform != "android":
            return path if os.path.exists(path) else None

        # Android content:// URI - copy to local temp
        try:
            from jnius import autoclass
            PythonActivity = autoclass("org.kivy.android.PythonActivity")
            Uri = autoclass("android.net.Uri")

            activity = PythonActivity.mActivity
            uri = Uri.parse(path)
            resolver = activity.getContentResolver()

            # Try to get _data column first
            try:
                cursor = resolver.query(uri, ["_data"], None, None, None)
                if cursor:
                    idx = cursor.getColumnIndex("_data")
                    if idx >= 0:
                        cursor.moveToFirst()
                        real_path = cursor.getString(idx)
                        cursor.close()
                        if real_path and os.path.exists(real_path):
                            return real_path
                    cursor.close()
            except Exception:
                pass

            # Fallback: copy stream to local temp
            istream = resolver.openInputStream(uri)
            self._temp_counter += 1
            ext = ".jpg"
            if prefix == "template":
                ext = ".xlsx"
            temp_path = os.path.join(
                self.temp_dir, f"{prefix}_{self._temp_counter}{ext}"
            )

            with open(temp_path, "wb") as f:
                buffer = bytearray(8192)
                while True:
                    # Java read() returns -1 at EOF or number of bytes
                    bytesRead = istream.read(buffer)
                    if bytesRead == -1:
                        break
                    f.write(buffer[:bytesRead])

            istream.close()
            Logger.info(f"QC Teazzi: Copied {prefix} to {temp_path} ({os.path.getsize(temp_path)} bytes)")
            return temp_path

        except Exception as e:
            Logger.error(f"QC Teazzi: URI resolve error: {e}")
            return None

    # ── PROSES ──
    def start_process(self):
        user = self.ids.inp_user.text.strip()
        toko = self.ids.inp_toko.text.strip()
        tanggal = self.ids.inp_tanggal.text.strip()

        if not user:
            self.ids.lbl_status.text = "[color=ff6b6b]Isi Nama Pengguna![/color]"
            return
        if not toko or not tanggal:
            self.ids.lbl_status.text = "[color=ff6b6b]Isi Nama Toko dan Tanggal![/color]"
            return
        if not self.photo_paths:
            self.ids.lbl_status.text = "[color=ff6b6b]Pilih foto terlebih dahulu![/color]"
            return

        template_text = self.ids.dd_template.text
        sheet = self.ids.dd_sheet.text
        layout = self.ids.dd_layout.text

        if not sheet:
            self.ids.lbl_status.text = "[color=ff6b6b]Pilih Target Sheet![/color]"
            return

        self.ids.btn_process.disabled = True
        self.ids.progress_bar.value = 0
        self.ids.result_box.opacity = 0
        self.ids.result_box.disabled = True
        self.ids.lbl_status.text = "Memproses..."

        threading.Thread(
            target=self._process_worker,
            args=(user, toko, tanggal, template_text, sheet, layout),
            daemon=True,
        ).start()

    def _process_worker(self, user, toko, tanggal, template_text, sheet, layout):
        try:
            Logger.info(f"QC Teazzi: Starting process - {user}, {toko}, {tanggal}, {layout}")
            Logger.info(f"QC Teazzi: Template={template_text}, Sheet={sheet}")
            Logger.info(f"QC Teazzi: Photos={len(self.photo_paths)}")

            # Load template
            template_bytes = None
            real_name = self.preset_loader.get_real_filename(template_text)

            if self._manual_template_bytes:
                template_bytes = self._manual_template_bytes
                Logger.info("QC Teazzi: Using manual template")
            elif real_name and self.preset_loader.is_cached(real_name):
                template_bytes = self.preset_loader.load_preset(real_name)
                Logger.info(f"QC Teazzi: Using cached preset {real_name}")

            if not template_bytes:
                Clock.schedule_once(
                    lambda dt: self._set_status("[color=ff6b6b]Template tidak ditemukan! Pilih template dulu.[/color]"),
                    0,
                )
                return

            # Check sheet exists
            from openpyxl import load_workbook
            template_bytes.seek(0)
            wb_check = load_workbook(template_bytes, read_only=True)
            Logger.info(f"QC Teazzi: Available sheets = {wb_check.sheetnames}")
            if sheet not in wb_check.sheetnames:
                wb_check.close()
                Clock.schedule_once(
                    lambda dt: self._set_status(
                        f"[color=ff6b6b]Sheet '{sheet}' tidak ditemukan! Tersedia: {wb_check.sheetnames}[/color]"
                    ),
                    0,
                )
                return
            wb_check.close()

            # Process
            processor = ExcelProcessor(template_bytes, sheet, layout)
            capacity = processor.get_capacity()
            Logger.info(f"QC Teazzi: Layout capacity = {capacity} cells")

            def on_progress(current, total):
                pct = int((current / total) * 100)
                Clock.schedule_once(
                    lambda dt: self._update_progress(pct, current, total), 0
                )

            output_bytes, success = processor.process(self.photo_paths, on_progress)
            Logger.info(f"QC Teazzi: Process complete, {success} photos inserted")

            # Save output
            filename = generate_filename(toko, tanggal)
            self.output_path = get_output_path(filename)

            with open(self.output_path, "wb") as f:
                f.write(output_bytes.getvalue())
            self.output_bytes = output_bytes

            Logger.info(f"QC Teazzi: Output saved to {self.output_path}")
            Logger.info(f"QC Teazzi: File size = {os.path.getsize(self.output_path)} bytes")

            # Traffic log (invisible)
            self.traffic.log(user, toko, tanggal, layout)

            Clock.schedule_once(
                lambda dt: self._on_process_done(success, filename), 0
            )

        except Exception as e:
            Logger.error(f"QC Teazzi: Process error: {e}")
            import traceback
            traceback.print_exc()
            Clock.schedule_once(
                lambda dt: self._set_status(f"[color=ff6b6b]Error: {str(e)[:100]}[/color]"), 0
            )
        finally:
            Clock.schedule_once(lambda dt: self._enable_button(), 0)

    def _update_progress(self, pct, current, total):
        self.ids.progress_bar.value = pct
        self.ids.lbl_status.text = f"Mengompres foto {current}/{total}..."

    def _on_process_done(self, count, filename):
        self.ids.lbl_status.text = (
            f"[color=6ee7b7]Berhasil! {count} foto disusun[/color]"
        )
        self.ids.lbl_output_path.text = f"📁 Download/QCTeazzi/output/{filename}"
        self.ids.result_box.opacity = 1
        self.ids.result_box.disabled = False

    def _set_status(self, text):
        self.ids.lbl_status.text = text

    def _enable_button(self):
        self.ids.btn_process.disabled = False

    # ── OUTPUT ACTIONS ──
    def share_file(self):
        if not self.output_path or not os.path.exists(self.output_path):
            self.ids.lbl_status.text = "[color=ff6b6b]File tidak ditemukan![/color]"
            return

        if platform == "android":
            try:
                from jnius import autoclass
                from android import mActivity

                Context = autoclass("android.content.Context")
                Intent = autoclass("android.content.Intent")
                FileProvider = autoclass("android.support.v4.content.FileProvider")
                File = autoclass("java.io.File")
                Uri = autoclass("android.net.Uri")

                activity = mActivity
                f = File(self.output_path)

                # Get authority
                authority = activity.getPackageName() + ".fileprovider"
                uri = FileProvider.getUriForFile(activity, authority, f)

                intent = Intent(Intent.ACTION_SEND)
                intent.setType(
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                intent.putExtra(Intent.EXTRA_STREAM, uri)
                intent.addFlags(Intent.FLAG_GRANT_READ_URI_PERMISSION)
                intent.addFlags(Intent.FLAG_ACTIVITY_NEW_TASK)

                chooser = Intent.createChooser(intent, "Share QC Report")
                chooser.setFlags(Intent.FLAG_ACTIVITY_NEW_TASK)
                activity.startActivity(chooser)

                Logger.info("QC Teazzi: Share intent launched")
            except Exception as e:
                Logger.error(f"QC Teazzi: Share error: {e}")
                # Fallback: try plyer
                try:
                    from plyer import share
                    share.share(
                        title="Share QC Report",
                        text="Laporan QC dari QC Teazzi",
                        filepath=self.output_path,
                    )
                except Exception as e2:
                    Logger.error(f"QC Teazzi: Plyer share error: {e2}")
                    self.ids.lbl_status.text = f"[color=ff6b6b]Share gagal: {e2}[/color]"
        else:
            self.ids.lbl_status.text = "[color=fcd34d]Share hanya tersedia di Android[/color]"

    def open_folder(self):
        if not self.output_path:
            return
        folder = os.path.dirname(self.output_path)

        if platform == "android":
            try:
                from jnius import autoclass
                from android import mActivity

                Intent = autoclass("android.content.Intent")
                Uri = autoclass("android.net.Uri")

                activity = mActivity
                uri = Uri.parse(f"content://com.android.externalstorage.documents/document/primary:Download%2FQCTeazzi%2Foutput")
                intent = Intent(Intent.ACTION_VIEW)
                intent.setDataAndType(uri, "vnd.android.document/directory")
                intent.setFlags(Intent.FLAG_ACTIVITY_NEW_TASK)

                try:
                    activity.startActivity(intent)
                except Exception:
                    # Fallback: open Download folder
                    uri2 = Uri.parse("content://com.android.externalstorage.documents/document/primary:Download")
                    intent2 = Intent(Intent.ACTION_VIEW)
                    intent2.setDataAndType(uri2, "vnd.android.document/directory")
                    intent2.setFlags(Intent.FLAG_ACTIVITY_NEW_TASK)
                    activity.startActivity(intent2)

                Logger.info("QC Teazzi: Folder opened")
            except Exception as e:
                Logger.error(f"QC Teazzi: Open folder error: {e}")
                self.ids.lbl_status.text = f"[color=fcd34d]Buka File Manager → Download → QCTeazzi → output[/color]"
        else:
            self.ids.lbl_status.text = f"[color=fcd34d]Folder: {folder}[/color]"

    def open_file(self):
        if not self.output_path or not os.path.exists(self.output_path):
            self.ids.lbl_status.text = "[color=ff6b6b]File tidak ditemukan![/color]"
            return

        if platform == "android":
            try:
                from jnius import autoclass
                from android import mActivity

                Context = autoclass("android.content.Context")
                Intent = autoclass("android.content.Intent")
                FileProvider = autoclass("android.support.v4.content.FileProvider")
                File = autoclass("java.io.File")
                Uri = autoclass("android.net.Uri")

                activity = mActivity
                f = File(self.output_path)

                authority = activity.getPackageName() + ".fileprovider"
                uri = FileProvider.getUriForFile(activity, authority, f)

                intent = Intent(Intent.ACTION_VIEW)
                intent.setDataAndType(
                    uri,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
                intent.addFlags(Intent.FLAG_GRANT_READ_URI_PERMISSION)
                intent.addFlags(Intent.FLAG_ACTIVITY_NEW_TASK)

                activity.startActivity(intent)
                Logger.info("QC Teazzi: File opened")
            except Exception as e:
                Logger.error(f"QC Teazzi: Open file error: {e}")
                self.ids.lbl_status.text = (
                    f"[color=fcd34d]File tersimpan di Download/QCTeazzi/output/[/color]"
                )
        else:
            os.startfile(self.output_path) if os.name == "nt" else os.system(
                f"xdg-open '{self.output_path}'"
            )
