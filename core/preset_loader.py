import os
from io import BytesIO
from openpyxl import load_workbook
from core.config import PRESETS_DIR, PRESET_FILES


class PresetLoader:
    def __init__(self):
        os.makedirs(PRESETS_DIR, exist_ok=True)

    def get_dropdown_options(self):
        options = []
        for filename in PRESET_FILES:
            cached = self.is_cached(filename)
            mark = " ✓" if cached else " ✗"
            options.append(f"{filename}{mark}")
        options.append("─────────────")
        options.append("Upload Manual")
        return options

    def list_cached(self):
        result = []
        for filename in PRESET_FILES:
            if self.is_cached(filename):
                result.append(filename)
        return result

    def load_preset(self, filename):
        path = os.path.join(PRESETS_DIR, filename)
        if os.path.exists(path):
            with open(path, "rb") as f:
                return BytesIO(f.read())
        return None

    def is_cached(self, filename):
        path = os.path.join(PRESETS_DIR, filename)
        return os.path.exists(path) and os.path.getsize(path) > 0

    def get_sheet_names(self, filename=None, file_bytes=None):
        try:
            if file_bytes:
                file_bytes.seek(0)
                wb = load_workbook(file_bytes, read_only=True)
            elif filename:
                path = os.path.join(PRESETS_DIR, filename)
                wb = load_workbook(path, read_only=True)
            else:
                return []
            names = wb.sheetnames
            wb.close()
            return names
        except Exception:
            return []

    def get_preset_path(self, filename):
        return os.path.join(PRESETS_DIR, filename)

    def get_real_filename(self, display_name):
        return display_name.replace(" ✓", "").replace(" ✗", "").strip()
